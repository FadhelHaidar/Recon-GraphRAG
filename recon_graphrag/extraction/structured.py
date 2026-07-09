"""Structured-data (row/record) ingestion: mapping spec, row conversion, sources.

Rows are plain dicts. A RowMapping declares how columns become entities,
relationships, chunk metadata, and which free-text columns get an LLM
extraction pass. The direct path emits the same GraphExtraction intermediate
the LLM emits, so the assembler/writer/finalize machinery is reused unchanged.

Direct entities use label-scoped canonical keys ("Product:Widget") so the same
name under two labels stays two nodes; direct and LLM-extracted entities merge
later via label-scoped entity resolution in the pipeline finalize step.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterator

from recon_graphrag.extraction.chunking import TextChunk
from recon_graphrag.extraction.schema import (
    GraphSchema,
    NodeType,
    PropertyDataType,
    PropertyType,
    RelationshipType,
)
from recon_graphrag.extraction.types import (
    ExtractedNode,
    ExtractedRelationship,
    GraphExtraction,
)


@dataclass(frozen=True)
class ColumnEntity:
    """Map row columns to one entity per row.

    ``identity_column``'s value becomes the entity name; the canonical key is
    label-scoped (``f"{label}:{value}"``). ``properties`` maps columns to
    property names; ``types`` maps property names to PropertyDataType (default
    STRING). ``description_template`` is ``str.format``-ed over the full row;
    when None, a compact "prop: value" summary of the mapped columns is used.
    """

    label: str
    identity_column: str
    properties: dict[str, str] = field(default_factory=dict)
    types: dict[str, PropertyDataType] = field(default_factory=dict)
    description_template: str | None = None

    def scoped_key(self, identity_value: str) -> str:
        return f"{self.label}:{identity_value}"


@dataclass(frozen=True)
class RowRelationship:
    """A relationship between two ColumnEntity labels within the same row."""

    source_label: str
    type: str
    target_label: str
    properties: dict[str, str] = field(default_factory=dict)


@dataclass(frozen=True)
class TextColumn:
    """A free-text column that triggers an LLM extraction pass for the row.

    Extracted entities are linked from the row's ``anchor_label`` entity via
    ``relationship`` edges (in addition to shared-chunk provenance).
    """

    column: str
    anchor_label: str | None = None
    relationship: str = "MENTIONS"


@dataclass(frozen=True)
class RowMapping:
    entities: list[ColumnEntity]
    relationships: list[RowRelationship] = field(default_factory=list)
    text_columns: list[TextColumn] = field(default_factory=list)
    record_id_column: str | None = None
    metadata_columns: list[str] = field(default_factory=list)

    def validate(self) -> None:
        if not self.entities:
            raise ValueError("RowMapping requires at least one ColumnEntity")

        labels = [e.label for e in self.entities]
        duplicates = {label for label in labels if labels.count(label) > 1}
        if duplicates:
            raise ValueError(f"Duplicate ColumnEntity labels: {sorted(duplicates)}")

        known = set(labels)
        for rel in self.relationships:
            for endpoint in (rel.source_label, rel.target_label):
                if endpoint not in known:
                    raise ValueError(
                        f"RowRelationship {rel.type!r} references unknown label "
                        f"{endpoint!r}; known labels: {sorted(known)}"
                    )
        for tc in self.text_columns:
            if tc.anchor_label is not None and tc.anchor_label not in known:
                raise ValueError(
                    f"TextColumn {tc.column!r} anchors to unknown label "
                    f"{tc.anchor_label!r}; known labels: {sorted(known)}"
                )

    def to_schema(self) -> GraphSchema:
        """Build a GraphSchema from the mapping (used for the LLM text pass)."""
        node_types = [
            NodeType(
                label=entity.label,
                properties=[
                    PropertyType(name=prop, type=entity.types.get(prop, "STRING"))
                    for prop in entity.properties.values()
                ],
            )
            for entity in self.entities
        ]
        relationship_labels = {rel.type for rel in self.relationships}
        relationship_labels.update(tc.relationship for tc in self.text_columns)
        return GraphSchema(
            node_types=node_types,
            relationship_types=[
                RelationshipType(label=label)
                for label in sorted(relationship_labels)
            ],
        )

    def resolution_context(self) -> dict[str, list[str]]:
        """Per-label typed property names for entity_resolution_context_properties."""
        return {
            entity.label: sorted(entity.properties.values())
            for entity in self.entities
            if entity.properties
        }


_TRUTHY = {"true", "yes", "y", "t", "1"}
_FALSY = {"false", "no", "n", "f", "0"}


def coerce(
    value: Any,
    ptype: PropertyDataType,
    *,
    column: str,
    row_index: int,
) -> Any:
    """Coerce a cell value to a PropertyDataType. None/empty string → None."""
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None

    try:
        if ptype == "STRING":
            return str(value)
        if ptype == "INTEGER":
            return int(value)
        if ptype == "FLOAT":
            return float(value)
        if ptype == "BOOLEAN":
            if isinstance(value, bool):
                return value
            lowered = str(value).lower()
            if lowered in _TRUTHY:
                return True
            if lowered in _FALSY:
                return False
            raise ValueError(f"not a boolean: {value!r}")
        if ptype == "DATE":
            if isinstance(value, datetime):
                return value.date().isoformat()
            if isinstance(value, date):
                return value.isoformat()
            return date.fromisoformat(str(value)).isoformat()
        if ptype == "DATETIME":
            if isinstance(value, datetime):
                return value.isoformat()
            return datetime.fromisoformat(str(value)).isoformat()
        if ptype == "LIST":
            if isinstance(value, (list, tuple)):
                return list(value)
            return [part.strip() for part in str(value).split(",") if part.strip()]
        raise ValueError(f"unsupported property type: {ptype}")
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"cannot coerce column {column!r} value {value!r} to {ptype} "
            f"(row {row_index}): {exc}"
        ) from None


def row_text(row: dict) -> str:
    """Serialize a row into readable "column: value" lines (chunk/citation text)."""
    lines = []
    for column, value in row.items():
        cleaned = _clean_cell(value)
        if cleaned is not None:
            lines.append(f"{column}: {cleaned}")
    return "\n".join(lines)


def rows_to_chunks_and_extractions(
    rows: list[dict],
    mapping: RowMapping,
    document_id: str,
) -> tuple[list[TextChunk], dict[str, GraphExtraction], dict[str, list[tuple[str, str, str]]]]:
    """Convert rows into per-row chunks and deterministic extractions.

    Returns (chunks, extractions_by_chunk_id, llm_rows) where llm_rows maps
    chunk ids that need an LLM text pass to their anchor specs:
    (anchor_scoped_key, anchor_identity_value, relationship_type). Anchor specs
    with an empty scoped key mean "extract but do not anchor".
    """
    chunks: list[TextChunk] = []
    extractions: dict[str, GraphExtraction] = {}
    llm_rows: dict[str, list[tuple[str, str, str]]] = {}

    for index, row in enumerate(rows):
        chunk_id = f"{document_id}:row:{index}"

        metadata: dict[str, Any] = {"row_index": index}
        record_id = (
            _clean_cell(row.get(mapping.record_id_column))
            if mapping.record_id_column
            else None
        )
        metadata["record_id"] = record_id if record_id is not None else index
        for column in mapping.metadata_columns:
            cleaned = _clean_cell(row.get(column))
            if cleaned is not None:
                metadata[column] = cleaned

        chunks.append(
            TextChunk(id=chunk_id, text=row_text(row), index=index, metadata=metadata)
        )

        nodes: list[ExtractedNode] = []
        keys_by_label: dict[str, tuple[str, str]] = {}  # label -> (scoped, identity)
        for entity in mapping.entities:
            identity = _clean_cell(row.get(entity.identity_column))
            if identity is None:
                continue
            identity = str(identity)
            scoped = entity.scoped_key(identity)
            keys_by_label[entity.label] = (scoped, identity)

            properties: dict[str, Any] = {"name": identity, "canonical_key": scoped}
            for column, prop_name in entity.properties.items():
                coerced = coerce(
                    row.get(column),
                    entity.types.get(prop_name, "STRING"),
                    column=column,
                    row_index=index,
                )
                if coerced is not None:
                    properties[prop_name] = coerced
            properties["description"] = _entity_description(entity, identity, properties, row)

            nodes.append(
                ExtractedNode(id=scoped, label=entity.label, properties=properties)
            )

        relationships: list[ExtractedRelationship] = []
        for rel in mapping.relationships:
            source = keys_by_label.get(rel.source_label)
            target = keys_by_label.get(rel.target_label)
            if source is None or target is None:
                continue
            rel_properties = {}
            for column, prop_name in rel.properties.items():
                cleaned = _clean_cell(row.get(column))
                if cleaned is not None:
                    rel_properties[prop_name] = cleaned
            relationships.append(
                ExtractedRelationship(
                    source_id=source[0],
                    target_id=target[0],
                    type=rel.type,
                    properties=rel_properties,
                )
            )

        extractions[chunk_id] = GraphExtraction(
            nodes=nodes, relationships=relationships
        )

        anchors: list[tuple[str, str, str]] = []
        for tc in mapping.text_columns:
            if _clean_cell(row.get(tc.column)) is None:
                continue
            anchor = (
                keys_by_label.get(tc.anchor_label) if tc.anchor_label else None
            )
            if anchor is not None:
                anchors.append((anchor[0], anchor[1], tc.relationship))
            else:
                anchors.append(("", "", tc.relationship))
        if anchors:
            llm_rows[chunk_id] = anchors

    return chunks, extractions, llm_rows


def _entity_description(
    entity: ColumnEntity, identity: str, properties: dict, row: dict
) -> str:
    if entity.description_template is not None:
        return entity.description_template.format(**row)
    parts = [
        f"{name}: {value}"
        for name, value in properties.items()
        if name not in ("name", "canonical_key")
    ]
    summary = f"{entity.label} {identity}"
    return f"{summary} — {'; '.join(parts)}" if parts else summary


def _clean_cell(value: Any) -> Any:
    """Normalize a cell: strip strings, map None/empty to None."""
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return value


def iter_csv(
    path: str | Path, encoding: str = "utf-8-sig", **fmtparams: Any
) -> Iterator[dict]:
    """Yield rows from a CSV file as dicts (header row = keys)."""
    with open(path, newline="", encoding=encoding) as handle:
        yield from csv.DictReader(handle, **fmtparams)


def iter_excel(path: str | Path, sheet: str | None = None) -> Iterator[dict]:
    """Yield rows from an Excel worksheet as dicts (first row = header)."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for Excel ingestion: "
            "pip install recon-graphrag[excel]"
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet] if sheet else workbook.active
        rows = worksheet.iter_rows(values_only=True)
        header = [str(cell) if cell is not None else "" for cell in next(rows, ())]
        for row in rows:
            yield dict(zip(header, row))
    finally:
        workbook.close()


def iter_sql(connection: Any, query: str, params: Any = None) -> Iterator[dict]:
    """Yield rows from any DB-API 2.0 connection as dicts (bring your own driver)."""
    cursor = connection.cursor()
    try:
        cursor.execute(query, params or ())
        columns = [description[0] for description in cursor.description]
        for row in cursor:
            yield dict(zip(columns, row))
    finally:
        cursor.close()


__all__ = [
    "ColumnEntity",
    "RowRelationship",
    "TextColumn",
    "RowMapping",
    "coerce",
    "row_text",
    "rows_to_chunks_and_extractions",
    "iter_csv",
    "iter_excel",
    "iter_sql",
]
